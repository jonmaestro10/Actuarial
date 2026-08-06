"""The workbook an audit file is kept in, and what it may not pretend to be.

RFC-047. Every incumbent meets the actuary in Excel on the way out, and so
does this — a run summary, per-variable aggregates, the assumption snapshot
and the RFC-033 parity report, written with :mod:`openpyxl` behind the
``[excel]`` extra. What makes this workbook different from the ones the
field already produces is what is stamped on it and what it refuses to do.

**Every sheet is stamped, because sheets get copied.** The run fingerprint,
the assumption digest and the results digest occupy the first rows of
*every* sheet, above the frozen pane. A provenance block on sheet one helps
whoever is holding sheet one; a tab pasted into a board pack takes its
provenance with it. :func:`read_stamps` reads them back, and a workbook
whose sheets disagree about which run they came from is a workbook this
module will not have written.

**A workbook is a readable artifact, not a bit-exact transport.** This is
the finding the module is built around, and it is measured rather than
assumed: openpyxl serialises a float at 16 significant digits, one short of
the 17 a float64 needs to round-trip, so ``123456789.123456789`` comes back
a bit different; Excel itself then parses only 15. Non-finite values are
worse — openpyxl writes ``NaN`` and ``±inf`` as *empty cells*, and a blank
cell in a cashflow column reads as zero to every human being who has ever
opened a spreadsheet. So:

- non-finite values are written as the text ``NaN`` / ``+Inf`` / ``-Inf``
  and **counted on the run-summary sheet**, never left blank;
- the summary sheet states the precision limit in the workbook itself,
  next to the digest of the run that holds the exact numbers;
- the bit-exact record of a run remains the registry digest and the
  Parquet warehouse of RFC-046. The workbook points at them; it does not
  replace them.

**It refuses what will not fit rather than truncating it.** An Excel grid
is 1,048,576 rows by 16,384 columns. A model-point detail sheet for a
100,000-policy block does not fit, and the writer says so, naming the
warehouse, rather than writing the first 16,383 policies under a heading
that says "detail".

**The digest is over the content, not the bytes.** ``.xlsx`` is a zip of
XML whose serialisation is openpyxl's business and whose entry timestamps
are the clock's; the *content* — the cells this module decided to write —
is ours. :attr:`WorkbookWrite.content_digest` fingerprints the cells, so it
survives an openpyxl upgrade that reorders an attribute. The file bytes are
made reproducible anyway, by pinning the document properties and the zip
timestamps, because an audit artifact that changes every time it is written
invites the question of what else changed.
"""

from __future__ import annotations

import io
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np

from engine.core.fingerprint import fingerprint
from engine.core.registry import ArtifactRecord, ArtifactRegistry, git_commit
from engine.core.snapshot import MAX_DEPTH, MAX_ITEMS, snapshot_rows

#: Excel's grid, which is a hard limit rather than a guideline.
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
#: Excel's sheet-name limit, and the characters it will not accept in one.
SHEET_NAME_LIMIT = 31
FORBIDDEN_IN_SHEET_NAME = set("[]:*?/\\")

#: Significant decimal digits openpyxl serialises a float to. A float64
#: needs 17 to round-trip; Excel itself parses 15. Measured, not assumed —
#: tests/test_excel_workbook.py holds a value that demonstrates the loss.
SIGNIFICANT_DIGITS = 16

#: Pinned so that the same content writes the same file. The real clock is
#: on the summary sheet and in the run record, which is where a reader
#: should be looking for it anyway.
PINNED_TIMESTAMP = datetime(1980, 1, 1, 0, 0, 0)
PINNED_ZIP_DATE = (1980, 1, 1, 0, 0, 0)

WORKBOOK_KIND = "workbook"

SUMMARY_SHEET = "Run summary"
AGGREGATE_SHEET = "Aggregates"
ASSUMPTION_SHEET = "Assumptions"
PARITY_SHEET = "Parity"

#: Rows the stamp occupies at the top of every sheet, blank separator
#: included. Content starts below it and the pane freezes there.
STAMP_ROWS = 5

NAN_TEXT = "NaN"
POS_INF_TEXT = "+Inf"
NEG_INF_TEXT = "-Inf"


class ExcelError(ValueError):
    """A workbook this module will not write.

    Raised where the alternative is a spreadsheet that looks authoritative
    and is not: a parity report belonging to a different run, an assumption
    snapshot that does not re-derive to the run's assumption digest, a
    detail block wider than the Excel grid.
    """


def _openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - exercised by the extra
        raise ImportError(
            "the Excel surface needs openpyxl: pip install -e '.[excel]'"
        ) from exc
    return openpyxl, Font, get_column_letter


# --------------------------------------------------------------------------
# Cells
# --------------------------------------------------------------------------

def _cell(value: Any) -> Any:
    """One value, in the form the workbook will carry it.

    Non-finite floats become text. openpyxl writes them as empty cells,
    and an empty cell in a claims column is read as zero by everybody. The
    text is ugly on purpose: it is meant to stop a reader, not to blend in.
    """
    if isinstance(value, (float, np.floating)):
        number = float(value)
        if math.isnan(number):
            return NAN_TEXT
        if math.isinf(number):
            return POS_INF_TEXT if number > 0 else NEG_INF_TEXT
        return number
    if isinstance(value, (bool, np.bool_)):
        return bool(value)
    if isinstance(value, (int, np.integer)):
        return int(value)
    if value is None or isinstance(value, str):
        return value
    return str(value)


def as_written(value: float) -> float:
    """The number a float64 becomes on the way through a workbook.

    openpyxl serialises at :data:`SIGNIFICANT_DIGITS`, so this is the value
    a reader will get back — offered as a function because "close enough"
    is not a thing a caller should have to estimate. Compare against it to
    know whether a cell disagrees with the engine because of the format or
    because of the model.

    Excel-the-application parses 15 significant digits rather than 16, so
    this is the *tightest* the workbook can be, not a guarantee of what a
    given spreadsheet will show.
    """
    number = float(value)
    if not math.isfinite(number):
        return number
    return float(f"{number:.{SIGNIFICANT_DIGITS}g}")


def _nonfinite(values: Iterable[Any]) -> int:
    return sum(1 for value in values
               if value in (NAN_TEXT, POS_INF_TEXT, NEG_INF_TEXT))


@dataclass
class _Sheet:
    """A sheet under construction: a name and the rows going onto it.

    Built in memory before anything touches openpyxl so that the content
    digest is over what was decided rather than over what a library did
    with it, and so that a refusal happens before a file exists.
    """

    name: str
    rows: list[list[Any]]
    widths: dict[int, int]
    bold_rows: set[int]

    def __init__(self, name: str):
        self.name = name
        self.rows = []
        self.widths = {}
        self.bold_rows = set()

    def append(self, *values: Any, bold: bool = False) -> None:
        row = [_cell(value) for value in values]
        if bold:
            self.bold_rows.add(len(self.rows))
        self.rows.append(row)

    def blank(self) -> None:
        self.rows.append([])

    def width(self, column: int, chars: int) -> None:
        self.widths[column] = max(self.widths.get(column, 0), chars)

    @property
    def n_nonfinite(self) -> int:
        return sum(_nonfinite(row) for row in self.rows)


# --------------------------------------------------------------------------
# The stamp
# --------------------------------------------------------------------------

@dataclass(frozen=True)
class Stamp:
    """The provenance block at the top of a sheet."""

    run_id: str
    assumptions_digest: str
    results_digest: str
    sheet: str

    def __fingerprint__(self):
        return {"run_id": self.run_id,
                "assumptions_digest": self.assumptions_digest,
                "results_digest": self.results_digest}


def _stamp(sheet: _Sheet, record, title: str) -> None:
    """Write the provenance block. Called for every sheet, without exception."""
    sheet.append("run fingerprint", record.run_id, bold=True)
    sheet.append("assumptions digest", record.assumptions_digest)
    sheet.append("results digest", record.results_digest)
    sheet.append(title, bold=True)
    sheet.blank()
    sheet.width(1, 26)
    sheet.width(2, 38)
    assert len(sheet.rows) == STAMP_ROWS


def read_stamps(path: Path | str) -> dict[str, Stamp]:
    """Read the provenance block back off every sheet of a workbook.

    The check a reviewer can run without this repo's help is the same one:
    the three digests are in the top-left corner of every tab. This does it
    in code so a test can assert that no sheet escaped the stamp.
    """
    openpyxl, _, _ = _openpyxl()
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        stamps = {}
        for sheet in book.worksheets:
            values = {}
            for row in sheet.iter_rows(min_row=1, max_row=3, max_col=2,
                                       values_only=True):
                if row and row[0]:
                    values[str(row[0])] = row[1]
            missing = {"run fingerprint", "assumptions digest",
                       "results digest"} - set(values)
            if missing:
                raise ExcelError(
                    f"sheet {sheet.title!r} is missing {sorted(missing)} from "
                    f"its stamp"
                )
            stamps[sheet.title] = Stamp(
                run_id=str(values["run fingerprint"]),
                assumptions_digest=str(values["assumptions digest"]),
                results_digest=str(values["results digest"]),
                sheet=sheet.title,
            )
        return stamps
    finally:
        book.close()


# --------------------------------------------------------------------------
# The assumption snapshot
# --------------------------------------------------------------------------

def assumption_rows(assumptions: Any, *, max_depth: int = MAX_DEPTH,
                    max_items: int = MAX_ITEMS) -> list[dict]:
    """Flatten an assumption set into ``(path, kind, value, digest)`` rows.

    :func:`~engine.core.snapshot.snapshot_rows` under another name, and
    deliberately not a second walker: the rows on the snapshot sheet are
    the rows RFC-048's assumption-diff route joins, so two surfaces that
    disagreed about what a basis contains would be worse than either.

    What the sheet gets out of it is that every row carries the digest of
    the subtree beneath it and the root row carries the run's
    ``assumptions_digest`` — which is what turns a snapshot sheet from a
    description into evidence.
    """
    return snapshot_rows(assumptions, max_depth=max_depth,
                         max_items=max_items)


# --------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------

def _summary_sheet(record, *, label: str | None, n_modelpoints: int,
                   variables: Sequence[str], scenario: Any,
                   sheets: Sequence[str]) -> _Sheet:
    sheet = _Sheet(SUMMARY_SHEET)
    _stamp(sheet, record, "Run summary")
    if label:
        sheet.append("label", label)
    sheet.append("model", f"{record.model_module}.{record.model_name}")
    sheet.append("executor", record.executor)
    sheet.append("engine version", record.engine_version)
    sheet.append("code version", record.code_version)
    sheet.append("run recorded at", record.created_at)
    sheet.blank()
    sheet.append("digest", "value", bold=True)
    sheet.append("run fingerprint", record.run_id)
    sheet.append("model source", record.model_source_digest)
    sheet.append("assumptions", record.assumptions_digest)
    sheet.append("model points", record.modelpoints_digest)
    sheet.append("scenarios", record.scenarios_digest)
    sheet.append("results", record.results_digest)
    sheet.blank()
    sheet.append("shape", "value", bold=True)
    sheet.append("model points", n_modelpoints)
    sheet.append("projection length", record.proj_len)
    sheet.append("scenarios", record.n_scenarios)
    sheet.append("scenario written", "—" if scenario is None else str(scenario))
    sheet.append("variables", len(variables))
    sheet.append("sheets", ", ".join(sheets))
    sheet.width(2, 44)
    return sheet


def _precision_note(sheet: _Sheet, n_nonfinite: int) -> None:
    """What this workbook is not, said in the workbook.

    A limit discovered by a reader is a limit that has already cost
    somebody an afternoon.
    """
    sheet.blank()
    sheet.append("what this workbook is", bold=True)
    sheet.append(
        "precision",
        f"values are written to {SIGNIFICANT_DIGITS} significant digits; a "
        f"float64 needs 17 to round-trip and Excel itself parses 15, so a "
        f"cell here may differ from the engine's number in the last bits",
    )
    sheet.append(
        "exact record",
        "the bit-exact results are the run's registry entry and its Parquet "
        "warehouse rows (RFC-046), both keyed by the run fingerprint above",
    )
    sheet.append(
        "non-finite values",
        f"{n_nonfinite} cell(s) written as {NAN_TEXT}/{POS_INF_TEXT}/"
        f"{NEG_INF_TEXT} text rather than left blank",
    )
    sheet.width(2, 44)


def _aggregate_sheet(record, result, variables: Sequence[str],
                     scenario: Any) -> _Sheet:
    """Per-variable aggregates by time step.

    The values are the *engine's* aggregate — ``result.aggregate`` — not a
    ``SUM()`` formula over a detail sheet. A spreadsheet recomputing the
    total would do it in Excel's op order and could disagree with the run in
    the last bits, leaving the workbook contradicting the digest printed at
    the top of it. The workbook reports the run; it does not re-derive it.
    """
    sheet = _Sheet(AGGREGATE_SHEET)
    _stamp(sheet, record, "Aggregate by time step (sum over model points)")
    sheet.append("t", *variables, bold=True)
    columns = [_aggregate(result, name, scenario) for name in variables]
    n_steps = len(columns[0]) if columns else 0
    for t in range(n_steps):
        sheet.append(t, *(column[t] for column in columns))
    for index in range(1, len(variables) + 2):
        sheet.width(index, 18)
    return sheet


def _aggregate(result, name: str, scenario: Any) -> list[float]:
    array = _array(result, name)
    if array.ndim == 3:
        array = _slice_scenario(array, scenario)
    return array.sum(axis=1).tolist()


def _array(result, name: str) -> np.ndarray:
    if hasattr(result, "array"):
        try:
            return np.asarray(result.array(name))
        except KeyError:
            raise ExcelError(f"the run carries no variable {name!r}") from None
    try:
        return np.array([mp[name] for mp in result.per_mp], dtype=float).T
    except KeyError:
        raise ExcelError(f"the run carries no variable {name!r}") from None


def _slice_scenario(array: np.ndarray, scenario: Any) -> np.ndarray:
    if scenario == "mean":
        return array.mean(axis=2)
    return array[:, :, int(scenario)]


def _detail_sheet(record, result, name: str, mp_ids: Sequence[Any],
                  scenario: Any, used: set[str]) -> _Sheet:
    array = _array(result, name)
    if array.ndim == 3:
        array = _slice_scenario(array, scenario)
    n_steps, n_mp = array.shape
    if n_mp + 1 > EXCEL_MAX_COLUMNS:
        raise ExcelError(
            f"a detail sheet for {name!r} needs {n_mp + 1:,} columns and an "
            f"Excel sheet holds {EXCEL_MAX_COLUMNS:,}; write the model-point "
            f"detail to the warehouse (engine.data.warehouse) and keep the "
            f"workbook to aggregates"
        )
    if n_steps + STAMP_ROWS + 1 > EXCEL_MAX_ROWS:
        raise ExcelError(
            f"a detail sheet for {name!r} needs {n_steps:,} rows and an Excel "
            f"sheet holds {EXCEL_MAX_ROWS:,}"
        )
    title = _sheet_name(f"Detail {name}", used)
    sheet = _Sheet(title)
    _stamp(sheet, record, f"{name} by model point")
    sheet.append("t", *[str(mp_id) for mp_id in mp_ids], bold=True)
    for t in range(n_steps):
        sheet.append(t, *array[t].tolist())
    sheet.width(1, 6)
    return sheet


def _sheet_name(proposed: str, used: set[str]) -> str:
    """An Excel-legal sheet name, refusing a collision rather than making one.

    Excel truncates at 31 characters, so two variables with a long shared
    prefix can end up asking for the same tab. Renaming one silently is how
    a reader ends up reading the wrong variable's numbers; this says so.
    """
    cleaned = "".join(" " if char in FORBIDDEN_IN_SHEET_NAME else char
                      for char in proposed).strip()
    name = cleaned[:SHEET_NAME_LIMIT]
    if name in used:
        raise ExcelError(
            f"sheet name {name!r} (from {proposed!r}) is already taken; Excel "
            f"truncates a sheet name at {SHEET_NAME_LIMIT} characters, so two "
            f"variables collided — write them to separate workbooks"
        )
    used.add(name)
    return name


def _assumption_sheet(record, assumptions: Any, **options) -> _Sheet:
    sheet = _Sheet(ASSUMPTION_SHEET)
    _stamp(sheet, record, "Assumption snapshot")
    if assumptions is None:
        # Present and honest beats absent. A missing sheet is read as "there
        # were no assumptions"; this says what is known and what is not.
        sheet.append(
            "note",
            "the assumption object was not passed to the writer, so only its "
            "digest is on record — it is on the stamp above and on the run "
            "summary",
        )
        sheet.width(2, 60)
        return sheet

    rows = assumption_rows(assumptions, **options)
    if rows[0]["digest"] != record.assumptions_digest:
        raise ExcelError(
            f"the assumption set passed to the writer digests to "
            f"{rows[0]['digest']} and the run was made on "
            f"{record.assumptions_digest}; a snapshot sheet describing a "
            f"different basis from the run is worse than no sheet"
        )
    sheet.append("path", "kind", "value", "digest", bold=True)
    for row in rows:
        sheet.append(row["path"], row["kind"], row["value"], row["digest"])
    sheet.blank()
    sheet.append(
        "note",
        "each row digests the subtree beneath it; the first row is the "
        "run's assumptions digest, so a basis that differs differs in a row",
    )
    sheet.width(1, 38)
    sheet.width(2, 20)
    sheet.width(4, 36)
    return sheet


def _parity_sheet(record, report) -> _Sheet:
    """The RFC-033 report, on a tab.

    The refusal is the interesting part: a report whose ``results_digest``
    names a different run does not go into this run's workbook. A
    reconciliation stapled to the wrong results is the exact document a
    reviewer would be entitled to be angry about.
    """
    if (report.results_digest is not None
            and report.results_digest != record.results_digest):
        raise ExcelError(
            f"the parity report reconciles results {report.results_digest} "
            f"and this workbook is for {record.results_digest}"
        )
    sheet = _Sheet(PARITY_SHEET)
    _stamp(sheet, record, "Parity report (RFC-033)")
    sheet.append("verdict",
                 "PARITY — every mapped cell within tolerance" if report.ok
                 else "DIFFERENCES FOUND", bold=True)
    if report.label:
        sheet.append("label", report.label)
    sheet.append("external digest", report.external_digest)
    sheet.append("external source", report.external_source)
    sheet.append("spec digest", report.spec_digest)
    sheet.append("external rows", report.n_external_rows)
    sheet.append("rows matched", report.n_matched_rows)
    sheet.append("rows unmatched", report.n_unmatched_rows)
    sheet.append("engine cells covered", report.coverage)
    sheet.blank()

    sheet.append("variable", "column", "compared", "within", "outside",
                 "max abs", "max rel", "worst model point", "worst t",
                 "tolerance", bold=True)
    for entry in report.variables:
        sheet.append(entry.variable, entry.column, entry.n_compared,
                     entry.n_within, entry.n_outside, entry.max_absolute,
                     entry.max_relative,
                     None if entry.worst_modelpoint is None
                     else str(entry.worst_modelpoint),
                     entry.worst_t, entry.tolerance.describe())

    outside = [entry for entry in report.variables if not entry.ok]
    if outside:
        sheet.blank()
        sheet.append("cells outside tolerance", bold=True)
        sheet.append("variable", "model point", "t", "engine", "external",
                     "abs", "rel", bold=True)
        for entry in outside:
            for cell in entry.deviations:
                if cell.within:
                    continue
                sheet.append(entry.variable, str(cell.modelpoint), cell.t,
                             cell.engine, cell.external, cell.absolute,
                             cell.relative)

    if report.unmapped_columns:
        sheet.blank()
        sheet.append("external columns mapped to nothing, therefore "
                     "reconciled by nobody", bold=True)
        for name in report.unmapped_columns:
            sheet.append(name)

    sheet.width(1, 24)
    sheet.width(2, 20)
    sheet.width(10, 34)
    return sheet


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------

@dataclass(frozen=True)
class WorkbookWrite:
    """What one workbook write put where, and what it digests to."""

    path: Path
    run_id: str
    sheets: tuple[str, ...]
    n_variables: int
    n_cells: int
    n_nonfinite: int
    content_digest: str

    def __fingerprint__(self):
        return {"run_id": self.run_id, "sheets": list(self.sheets),
                "n_variables": self.n_variables, "n_cells": self.n_cells,
                "n_nonfinite": self.n_nonfinite,
                "content_digest": self.content_digest}


def write_workbook(path: Path | str, result, record, *,
                   assumptions: Any = None, parity=None,
                   variables: Sequence[str] | None = None,
                   detail: Sequence[str] = (), scenario: Any = None,
                   label: str | None = None,
                   max_depth: int = MAX_DEPTH,
                   max_items: int = MAX_ITEMS) -> WorkbookWrite:
    """Write one registered run to an ``.xlsx`` audit workbook.

    ``record`` is a :class:`~engine.core.registry.RunRecord`, for the same
    reason the warehouse takes one: provenance a caller assembled by hand is
    not provenance. ``assumptions`` is the object the run was made with —
    supply it and the snapshot sheet is written, and refused if it does not
    digest to the run's ``assumptions_digest``. ``parity`` is an RFC-033
    :class:`~engine.parity.diff.ParityReport`.

    A stochastic run must name a ``scenario`` — an index or ``"mean"`` —
    because a scenario silently picked is a number nobody can reproduce.
    """
    openpyxl, Font, get_column_letter = _openpyxl()
    names = list(variables if variables is not None else record.outputs)
    if not names:
        raise ExcelError(f"run {record.run_id} names no variables to write")
    for name in names:
        _array(result, name)  # raises for a variable the run does not carry

    probe = _array(result, names[0])
    if probe.ndim == 3:
        if scenario is None:
            raise ExcelError(
                f"run {record.run_id} is stochastic ({probe.shape[2]:,} "
                f"scenarios): name the scenario to write — an index, or "
                f"'mean' for the scenario mean"
            )
        if scenario != "mean" and not 0 <= int(scenario) < probe.shape[2]:
            raise ExcelError(
                f"scenario {scenario} outside 0..{probe.shape[2] - 1}"
            )
    elif scenario is not None:
        raise ExcelError(
            "this run has no scenario axis, so there is no scenario to pick"
        )

    unknown = [name for name in detail if name not in names]
    if unknown:
        raise ExcelError(
            f"detail asked for {unknown}, which the workbook is not writing; "
            f"it is writing {names}"
        )

    used: set[str] = {SUMMARY_SHEET, AGGREGATE_SHEET, ASSUMPTION_SHEET}
    body: list[_Sheet] = [
        _aggregate_sheet(record, result, names, scenario),
        _assumption_sheet(record, assumptions, max_depth=max_depth,
                          max_items=max_items),
    ]
    for name in detail:
        body.append(_detail_sheet(record, result, name, result.mp_ids,
                                  scenario, used))
    if parity is not None:
        used.add(PARITY_SHEET)
        body.append(_parity_sheet(record, parity))

    sheet_names = [SUMMARY_SHEET] + [sheet.name for sheet in body]
    summary = _summary_sheet(record, label=label,
                             n_modelpoints=len(result.mp_ids),
                             variables=names, scenario=scenario,
                             sheets=sheet_names)
    _precision_note(summary, sum(sheet.n_nonfinite for sheet in body))
    sheets = [summary] + body

    book = openpyxl.Workbook()
    book.remove(book.active)
    for built in sheets:
        worksheet = book.create_sheet(built.name)
        for index, row in enumerate(built.rows):
            worksheet.append(row)
            if index in built.bold_rows:
                for cell in worksheet[index + 1]:
                    cell.font = Font(bold=True)
        worksheet.freeze_panes = f"A{STAMP_ROWS + 1}"
        for column, chars in built.widths.items():
            worksheet.column_dimensions[get_column_letter(column)].width = chars
    book.properties.creator = "actuarial-engine"
    book.properties.created = PINNED_TIMESTAMP
    book.properties.modified = PINNED_TIMESTAMP
    book.properties.title = f"Run {record.run_id}"

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    _save_reproducibly(book, target)

    content = {"sheets": [{"name": sheet.name, "rows": sheet.rows}
                          for sheet in sheets]}
    return WorkbookWrite(
        path=target, run_id=record.run_id,
        sheets=tuple(sheet.name for sheet in sheets), n_variables=len(names),
        n_cells=sum(len(row) for sheet in sheets for row in sheet.rows),
        n_nonfinite=sum(sheet.n_nonfinite for sheet in sheets),
        content_digest=fingerprint(content),
    )


def _save_reproducibly(book, path: Path) -> None:
    """Save with the zip timestamps pinned, so identical content is one file.

    openpyxl stamps every zip entry with the wall clock, and overwrites the
    document's ``modified`` property at save time whatever the caller set,
    which makes two writes of the same workbook two different files. An
    audit artifact that changes when nothing changed trains its reader to
    ignore the fact that it changed.
    """
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    pinned_iso = PINNED_TIMESTAMP.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    with zipfile.ZipFile(buffer) as source:
        entries = []
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + pinned_iso + rb"\g<2>", payload,
                )
            entries.append((info, payload))
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as out:
        for info, payload in entries:
            pinned = zipfile.ZipInfo(info.filename, date_time=PINNED_ZIP_DATE)
            pinned.compress_type = info.compress_type
            pinned.external_attr = info.external_attr
            out.writestr(pinned, payload)


# --------------------------------------------------------------------------
# The registry
# --------------------------------------------------------------------------

def workbook_artifact(written: WorkbookWrite, *,
                      code_version: str | None = None) -> ArtifactRecord:
    """The registry record for a workbook.

    ``artifact_id`` digests what was asked for — the run and the sheets —
    and ``content_digest`` the cells that came out, so the same run written
    the same way twice is one artifact and a workbook that changed without
    its inputs changing is refused by the registry.
    """
    inputs: dict[str, Any] = {
        "kind": WORKBOOK_KIND,
        "run_id": written.run_id,
        "sheets": list(written.sheets),
        "n_variables": written.n_variables,
    }
    return ArtifactRecord(
        artifact_id=fingerprint(inputs),
        kind=WORKBOOK_KIND,
        content_digest=written.content_digest,
        inputs=inputs,
        label=str(written.path.name),
        ok=True,
        code_version=code_version if code_version is not None else git_commit(),
    )


def record_workbook(written: WorkbookWrite, registry: ArtifactRegistry, *,
                    code_version: str | None = None) -> ArtifactRecord:
    """Record a workbook in ``registry`` and return its record."""
    return registry.add(workbook_artifact(written, code_version=code_version))

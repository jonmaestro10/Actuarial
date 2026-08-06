"""The workbook, and what it is not allowed to pretend to be.

RFC-047. The Excel surface is where a run meets the people who will act on
it, and a spreadsheet is the easiest place in the stack to state something
false with a straight face. So the suite is mostly about refusals and about
limits:

- **every sheet carries the run's three digests**, because tabs get copied
  out of workbooks and a provenance block on sheet one travels with sheet
  one only;
- **the aggregates are the engine's**, cell for cell, rather than something
  a spreadsheet recomputed;
- **the assumption snapshot re-derives to the run's assumption digest**, and
  a snapshot of some other basis is refused rather than written;
- **a parity report belonging to another run is refused**, which is the one
  document a reviewer would be right to be angry about;
- **non-finite values are never blank cells**, because openpyxl writes them
  as blanks and a blank in a claims column reads as zero;
- **the precision loss is asserted rather than assumed** — the test holds
  the documented limit up, so a future openpyxl that fixes it makes the
  documentation, not the code, the thing that is wrong;
- **what will not fit is refused rather than truncated**, at the Excel grid
  and at the 31-character sheet-name limit.
"""

import dataclasses
import hashlib

import numpy as np
import pytest

pytest.importorskip("openpyxl", reason="needs the [excel] extra")

import openpyxl  # noqa: E402

from engine.core.fingerprint import fingerprint  # noqa: E402
from engine.core.registry import (  # noqa: E402
    ArtifactConflictError,
    ArtifactRegistry,
    record_run,
)
from engine.core.results import ArrayRunResult, StochasticRunResult  # noqa: E402
from engine.data.assumptions import Assumptions, MortalityTable  # noqa: E402
from engine.data.modelpoints import ModelPoint  # noqa: E402
from engine.data.scenarios import ScenarioSet  # noqa: E402
from engine.excel.workbook import (  # noqa: E402
    AGGREGATE_SHEET,
    ASSUMPTION_SHEET,
    EXCEL_MAX_COLUMNS,
    PARITY_SHEET,
    SIGNIFICANT_DIGITS,
    STAMP_ROWS,
    SUMMARY_SHEET,
    ExcelError,
    as_written,
    assumption_rows,
    read_stamps,
    record_workbook,
    write_workbook,
)
from engine.library.term_life import TermLife  # noqa: E402
from engine.library.unit_linked import UnitLinkedGMDB  # noqa: E402
from engine.parity.diff import ExternalTable, ParitySpec, diff  # noqa: E402

QX = {age: min(0.0006 * 1.095 ** (age - 20), 0.5) for age in range(20, 111)}
ASSUMPTIONS = Assumptions(mortality=MortalityTable(QX), lapse=0.04,
                          interest=0.03, expense_per_policy=50.0)
POINTS = [
    ModelPoint(id="T1", age_at_entry=45, term_years=20, sum_assured=250_000.0,
               annual_premium=1_100.0, init_pols=1),
    ModelPoint(id="T2", age_at_entry=55, term_years=10, sum_assured=100_000.0,
               annual_premium=900.0, init_pols=1),
]
PROJ_LEN = 20
OUTPUTS = ["pols_if", "claims", "premiums"]


@pytest.fixture(scope="module")
def run():
    return record_run(TermLife, POINTS, ASSUMPTIONS, PROJ_LEN, outputs=OUTPUTS)


@pytest.fixture
def workbook(tmp_path, run):
    result, record = run
    written = write_workbook(tmp_path / "audit.xlsx", result, record,
                             assumptions=ASSUMPTIONS, label="year-end")
    return openpyxl.load_workbook(written.path), written, result, record


def _rows(sheet, min_row=1, max_col=12):
    return [row for row in sheet.iter_rows(min_row=min_row, max_col=max_col,
                                           values_only=True)]


def _lookup(sheet, key):
    """First value in column B whose column A is ``key``."""
    for row in _rows(sheet, max_col=2):
        if row[0] == key:
            return row[1]
    raise KeyError(key)


# --------------------------------------------------------------------------
# The stamp
# --------------------------------------------------------------------------

def test_every_sheet_carries_the_run_and_assumption_digests(tmp_path, run):
    """The claim the whole module rests on: a tab pasted into a board pack
    takes its provenance with it, so the stamp is on all of them and not
    only on the summary."""
    result, record = run
    external = _external_from(result, record, OUTPUTS)
    report = diff(ParitySpec.from_results(
        result, external, {name: name for name in OUTPUTS}),
        results_digest=record.results_digest)
    written = write_workbook(tmp_path / "audit.xlsx", result, record,
                             assumptions=ASSUMPTIONS, parity=report,
                             detail=["claims"])

    stamps = read_stamps(written.path)
    assert set(stamps) == set(written.sheets)
    assert len(stamps) == 5
    for stamp in stamps.values():
        assert stamp.run_id == record.run_id
        assert stamp.assumptions_digest == record.assumptions_digest
        assert stamp.results_digest == record.results_digest


def test_the_stamp_sits_above_the_frozen_pane_so_scrolling_cannot_lose_it(
        workbook):
    book, _, _, record = workbook
    for sheet in book.worksheets:
        assert sheet.freeze_panes == f"A{STAMP_ROWS + 1}"
        assert sheet["B1"].value == record.run_id


def test_a_sheet_without_a_stamp_is_reported_rather_than_read(tmp_path):
    """``read_stamps`` is the check a reviewer runs; it has to fail on a
    workbook somebody else wrote, not quietly return an empty dict."""
    book = openpyxl.Workbook()
    book.active.title = "Numbers"
    book.active["A1"] = "run fingerprint"
    book.active["B1"] = "deadbeef"
    book.save(tmp_path / "unstamped.xlsx")
    with pytest.raises(ExcelError, match="missing.*assumptions digest"):
        read_stamps(tmp_path / "unstamped.xlsx")


# --------------------------------------------------------------------------
# The numbers
# --------------------------------------------------------------------------

def test_the_aggregates_are_the_engines_own_reduction(workbook):
    """Not a SUM() over a detail sheet: a spreadsheet recomputing the total
    would use Excel's op order and could disagree with the digest stamped at
    the top of the same sheet.

    Held to ``as_written`` — exact equality with the engine's float64 down
    to the format's own limit, and no looser. A tolerance here would hide
    the difference between a serialisation artefact and a wrong number."""
    book, _, result, _ = workbook
    sheet = book[AGGREGATE_SHEET]
    header = _rows(sheet, min_row=STAMP_ROWS + 1)[0]
    assert header[0] == "t"
    assert list(header[1:1 + len(OUTPUTS)]) == OUTPUTS

    body = _rows(sheet, min_row=STAMP_ROWS + 2)
    assert len(body) == PROJ_LEN + 1
    for index, name in enumerate(OUTPUTS, start=1):
        column = [row[index] for row in body]
        assert column == [as_written(v) for v in result.aggregate(name)]
    assert [row[0] for row in body] == list(range(PROJ_LEN + 1))


def test_a_detail_sheet_carries_every_model_point_column(tmp_path, run):
    result, record = run
    written = write_workbook(tmp_path / "d.xlsx", result, record,
                             detail=["claims"])
    book = openpyxl.load_workbook(written.path)
    sheet = book["Detail claims"]
    header = _rows(sheet, min_row=STAMP_ROWS + 1)[0]
    assert list(header[:3]) == ["t", "T1", "T2"]
    body = _rows(sheet, min_row=STAMP_ROWS + 2)
    assert [row[1] for row in body] == \
        [as_written(v) for v in result.array("claims")[:, 0].tolist()]


def test_the_summary_names_the_run_the_shape_and_the_sheets(workbook):
    book, written, _, record = workbook
    sheet = book[SUMMARY_SHEET]
    assert _lookup(sheet, "label") == "year-end"
    assert _lookup(sheet, "model") == "engine.library.term_life.TermLife"
    assert _lookup(sheet, "executor") == record.executor
    assert _lookup(sheet, "projection length") == PROJ_LEN
    assert _lookup(sheet, "model points") == record.modelpoints_digest
    assert _lookup(sheet, "sheets") == ", ".join(written.sheets)


# --------------------------------------------------------------------------
# What a workbook is not: precision, and blanks
# --------------------------------------------------------------------------

def test_the_workbook_is_not_a_bit_exact_transport_and_the_limit_is_real():
    """The documented limit, asserted. openpyxl serialises a float at
    16 significant digits and a float64 needs 17, so this pair — distinct
    doubles — becomes one number on the way through. If a future openpyxl
    fixes that, this test fails and RFC-047's precision claim gets revisited
    rather than quietly becoming false."""
    import io

    a = 123456789.123456789
    b = np.nextafter(a, np.inf)
    assert a != b

    book = openpyxl.Workbook()
    book.active.append([a, b])
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    back = openpyxl.load_workbook(buffer).active
    assert back["A1"].value == back["B1"].value
    assert f"{a:.{SIGNIFICANT_DIGITS}g}" == f"{b:.{SIGNIFICANT_DIGITS}g}"


def test_the_summary_states_the_precision_limit_in_the_workbook_itself(
        workbook):
    """A limit a reader discovers has already cost somebody an afternoon."""
    book, _, _, _ = workbook
    note = _lookup(book[SUMMARY_SHEET], "precision")
    assert f"{SIGNIFICANT_DIGITS} significant digits" in note
    assert "17" in note and "15" in note
    assert "warehouse" in _lookup(book[SUMMARY_SHEET], "exact record")


def test_non_finite_values_are_text_and_counted_never_blank(tmp_path, run):
    """openpyxl writes NaN and ±inf as empty cells, and an empty cell in a
    claims column is read as zero by every human being alive."""
    _, record = run
    array = np.zeros((3, 2))
    array[0, 0] = np.nan
    array[1, 0] = np.inf
    array[2, 0] = -np.inf
    result = ArrayRunResult({"claims": array}, ["T1", "T2"])
    stub = dataclasses.replace(record, outputs=("claims",), proj_len=2)

    written = write_workbook(tmp_path / "nf.xlsx", result, stub,
                             detail=["claims"])
    assert written.n_nonfinite == 6        # three on each of two sheets
    book = openpyxl.load_workbook(written.path)
    column = [row[1] for row in _rows(book[AGGREGATE_SHEET],
                                      min_row=STAMP_ROWS + 2)]
    assert column == ["NaN", "+Inf", "-Inf"]
    assert _lookup(book[SUMMARY_SHEET], "non-finite values").startswith("6 ")


# --------------------------------------------------------------------------
# The assumption snapshot
# --------------------------------------------------------------------------

def test_the_snapshot_re_derives_to_the_runs_assumption_digest(workbook):
    """What turns a description into evidence: the first row of the sheet is
    the digest the registry recorded for the run."""
    book, _, _, record = workbook
    body = _rows(book[ASSUMPTION_SHEET], min_row=STAMP_ROWS + 2, max_col=4)
    header = _rows(book[ASSUMPTION_SHEET], min_row=STAMP_ROWS + 1,
                   max_col=4)[0]
    assert list(header) == ["path", "kind", "value", "digest"]
    assert body[0][0] == "Assumptions"
    assert body[0][3] == record.assumptions_digest
    paths = {row[0] for row in body if row[0]}
    assert "interest" in paths and "mortality" in paths


def test_a_changed_component_changes_its_own_rows_and_not_the_others():
    """The reason the sheet carries a digest per row rather than one for the
    set: a reviewer learns *which* component moved."""
    before = {row["path"]: row["digest"]
              for row in assumption_rows(ASSUMPTIONS)}
    changed = Assumptions(mortality=MortalityTable(QX), lapse=0.05,
                          interest=0.03, expense_per_policy=50.0)
    after = {row["path"]: row["digest"] for row in assumption_rows(changed)}

    moved = {path for path in before if before[path] != after.get(path)}
    assert "Assumptions" in moved                     # the root always moves
    assert any("lapse" in path for path in moved)
    assert before["interest"] == after["interest"]
    assert before["mortality"] == after["mortality"]


def test_a_snapshot_of_a_different_basis_is_refused(tmp_path, run):
    """The failure this exists to prevent: a workbook whose assumption tab
    describes a basis the run was not made on."""
    result, record = run
    other = Assumptions(mortality=MortalityTable(QX), lapse=0.10,
                        interest=0.03, expense_per_policy=50.0)
    with pytest.raises(ExcelError, match="worse than no sheet"):
        write_workbook(tmp_path / "x.xlsx", result, record, assumptions=other)


def test_an_unsupplied_assumption_set_leaves_a_sheet_that_says_so(tmp_path,
                                                                  run):
    """A missing sheet reads as "there were no assumptions"; a sheet that
    states what is known and what is not reads as the truth."""
    result, record = run
    written = write_workbook(tmp_path / "no-assumptions.xlsx", result, record)
    assert ASSUMPTION_SHEET in written.sheets
    book = openpyxl.load_workbook(written.path)
    assert "only its digest is on record" in _lookup(book[ASSUMPTION_SHEET],
                                                     "note")


def test_a_summarised_component_still_carries_the_digest_that_re_derives():
    """A snapshot sheet nobody can read is not a control, so deep and wide
    subtrees collapse to one row — but never to none, and never to a row
    without its digest."""
    shallow = assumption_rows(ASSUMPTIONS, max_depth=1)
    by_path = {row["path"]: row for row in shallow}
    assert by_path["mortality"]["digest"] == fingerprint(ASSUMPTIONS.mortality)
    assert not any(path.startswith("mortality.") for path in by_path)

    # A container wider than ``max_items`` collapses the same way: the root
    # alone, still digesting to what the registry recorded for the run.
    narrow = assumption_rows(ASSUMPTIONS, max_items=2)
    assert len(narrow) == 1
    assert narrow[0]["digest"] == fingerprint(ASSUMPTIONS)


# --------------------------------------------------------------------------
# The parity sheet
# --------------------------------------------------------------------------

def _external_from(result, record, names, perturb=None):
    """The engine's own numbers as somebody else's extract."""
    rows = []
    for column, mp_id in enumerate(result.mp_ids):
        for t in range(record.proj_len + 1):
            row = {"modelpoint_id": mp_id, "t": t}
            for name in names:
                value = float(result.array(name)[t, column])
                if perturb is not None and (mp_id, t, name) in perturb:
                    value = perturb[(mp_id, t, name)]
                row[name] = value
            rows.append(row)
    return ExternalTable.from_rows(rows, source="incumbent.csv")


def test_the_parity_sheet_carries_the_verdict_and_the_variable_table(tmp_path,
                                                                     run):
    result, record = run
    external = _external_from(result, record, OUTPUTS)
    report = diff(ParitySpec.from_results(
        result, external, {name: name for name in OUTPUTS}, label="Prophet"),
        results_digest=record.results_digest)
    assert report.ok

    written = write_workbook(tmp_path / "p.xlsx", result, record,
                             parity=report)
    sheet = openpyxl.load_workbook(written.path)[PARITY_SHEET]
    assert _lookup(sheet, "verdict").startswith("PARITY")
    assert _lookup(sheet, "external digest") == report.external_digest
    assert _lookup(sheet, "external source") == "incumbent.csv"
    names = {row[0] for row in _rows(sheet, min_row=STAMP_ROWS + 1)}
    assert set(OUTPUTS) <= names


def test_a_failing_reconciliation_names_the_cells_on_the_sheet(tmp_path, run):
    """A workbook that reported only "differences found" would be a workbook
    whose reader has to go and ask for the real report."""
    result, record = run
    external = _external_from(result, record, OUTPUTS,
                              perturb={("T1", 3, "claims"): 1.0})
    report = diff(ParitySpec.from_results(
        result, external, {name: name for name in OUTPUTS}),
        results_digest=record.results_digest)
    assert not report.ok

    written = write_workbook(tmp_path / "bad.xlsx", result, record,
                             parity=report)
    sheet = openpyxl.load_workbook(written.path)[PARITY_SHEET]
    assert _lookup(sheet, "verdict") == "DIFFERENCES FOUND"
    cells = [row for row in _rows(sheet, min_row=STAMP_ROWS + 1)
             if row[0] == "claims" and row[1] == "T1" and row[2] == 3]
    assert cells and cells[0][4] == 1.0


def test_a_parity_report_for_another_run_is_refused(tmp_path, run):
    """The document a reviewer would be entitled to be angry about: a
    reconciliation stapled to results it did not reconcile."""
    result, record = run
    external = _external_from(result, record, OUTPUTS)
    report = diff(ParitySpec.from_results(
        result, external, {name: name for name in OUTPUTS}),
        results_digest="0" * 32)
    with pytest.raises(ExcelError, match="reconciles results 0{32}"):
        write_workbook(tmp_path / "x.xlsx", result, record, parity=report)


def test_a_report_that_names_no_run_is_accepted(tmp_path, run):
    """A reconciliation of bare arrays does not claim to be another run's,
    so it is not refused — the refusal is for a mismatch, not for silence."""
    result, record = run
    external = _external_from(result, record, OUTPUTS)
    report = diff(ParitySpec.from_results(
        result, external, {name: name for name in OUTPUTS}))
    assert report.results_digest is None
    written = write_workbook(tmp_path / "q.xlsx", result, record,
                             parity=report)
    assert PARITY_SHEET in written.sheets


# --------------------------------------------------------------------------
# Refusals: scenarios, variables, and the Excel grid
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def stochastic():
    points = [ModelPoint(id="U1", age_at_entry=55, term_years=10,
                         premium=50_000.0, gmdb_guarantee=50_000.0,
                         init_pols=1)]
    scenarios = ScenarioSet.lognormal(n_scenarios=4, horizon=11, drift=0.03,
                                      vol=0.15, seed=7)
    assumptions = Assumptions(mortality=MortalityTable(QX), lapse=0.02,
                              interest=0.03, amc=0.01)
    return record_run(UnitLinkedGMDB, points, assumptions, 10,
                      outputs=["fund_eoy", "pols_if"], scenarios=scenarios)


def test_a_stochastic_run_refuses_to_pick_a_scenario_for_you(tmp_path,
                                                             stochastic):
    result, record = stochastic
    with pytest.raises(ExcelError, match="name the scenario"):
        write_workbook(tmp_path / "s.xlsx", result, record)
    with pytest.raises(ExcelError, match=r"scenario 9 outside 0\.\.3"):
        write_workbook(tmp_path / "s.xlsx", result, record, scenario=9)


def test_a_named_scenario_and_the_mean_both_write(tmp_path, stochastic):
    result, record = stochastic
    one = write_workbook(tmp_path / "one.xlsx", result, record, scenario=2)
    mean = write_workbook(tmp_path / "mean.xlsx", result, record,
                          scenario="mean")
    assert one.content_digest != mean.content_digest

    sheet = openpyxl.load_workbook(one.path)[AGGREGATE_SHEET]
    body = _rows(sheet, min_row=STAMP_ROWS + 2)
    expected = result.array("fund_eoy")[:, :, 2].sum(axis=1).tolist()
    assert [row[1] for row in body] == [as_written(v) for v in expected]
    assert _lookup(openpyxl.load_workbook(mean.path)[SUMMARY_SHEET],
                   "scenario written") == "mean"


def test_a_deterministic_run_refuses_a_scenario_selection(tmp_path, run):
    result, record = run
    with pytest.raises(ExcelError, match="no scenario axis"):
        write_workbook(tmp_path / "x.xlsx", result, record, scenario=0)


def test_a_variable_the_run_does_not_carry_is_named(tmp_path, run):
    result, record = run
    with pytest.raises(ExcelError, match="no variable 'surrenders'"):
        write_workbook(tmp_path / "x.xlsx", result, record,
                       variables=["claims", "surrenders"])
    with pytest.raises(ExcelError, match="detail asked for"):
        write_workbook(tmp_path / "x.xlsx", result, record,
                       variables=["claims"], detail=["premiums"])


def test_a_block_wider_than_the_excel_grid_is_refused_not_truncated(tmp_path,
                                                                    run):
    """16,383 policies under a heading that says "detail" is the failure
    mode; the message names the warehouse instead."""
    _, record = run
    n_mp = EXCEL_MAX_COLUMNS + 1
    result = ArrayRunResult({"claims": np.zeros((2, n_mp))},
                            [f"P{i}" for i in range(n_mp)])
    stub = dataclasses.replace(record, outputs=("claims",), proj_len=1)
    with pytest.raises(ExcelError, match="engine.data.warehouse"):
        write_workbook(tmp_path / "wide.xlsx", result, stub, detail=["claims"])
    # The aggregate of the same block is one column, so it still writes.
    assert write_workbook(tmp_path / "agg.xlsx", result, stub).n_variables == 1


def test_two_variables_colliding_at_the_sheet_name_limit_are_refused(tmp_path,
                                                                     run):
    """Excel truncates a sheet name at 31 characters. Renaming one silently
    is how a reader ends up reading the wrong variable's numbers."""
    _, record = run
    names = ("present_value_of_future_profits_gross",
             "present_value_of_future_profits_net")
    result = ArrayRunResult({name: np.zeros((2, 1)) for name in names}, ["P1"])
    stub = dataclasses.replace(record, outputs=names, proj_len=1)
    with pytest.raises(ExcelError, match="already taken"):
        write_workbook(tmp_path / "c.xlsx", result, stub, detail=list(names))


def test_a_run_with_no_outputs_is_refused(tmp_path, run):
    result, record = run
    stub = dataclasses.replace(record, outputs=())
    with pytest.raises(ExcelError, match="no variables to write"):
        write_workbook(tmp_path / "x.xlsx", result, stub)


# --------------------------------------------------------------------------
# The artifact
# --------------------------------------------------------------------------

def test_the_same_run_writes_the_same_file_byte_for_byte(tmp_path, run):
    """An audit artifact that changes when nothing changed trains its reader
    to ignore the fact that it changed. openpyxl stamps the zip entries and
    the document properties with the clock; the writer pins both."""
    result, record = run
    first = write_workbook(tmp_path / "a.xlsx", result, record,
                           assumptions=ASSUMPTIONS)
    second = write_workbook(tmp_path / "b.xlsx", result, record,
                            assumptions=ASSUMPTIONS)
    assert first.content_digest == second.content_digest
    digest = lambda path: hashlib.sha256(path.read_bytes()).hexdigest()  # noqa: E731
    assert digest(first.path) == digest(second.path)


def test_the_registry_records_one_artifact_and_refuses_a_second_answer(
        tmp_path, run):
    result, record = run
    registry = ArtifactRegistry()
    written = write_workbook(tmp_path / "a.xlsx", result, record,
                             assumptions=ASSUMPTIONS)
    stored = record_workbook(written, registry)
    assert stored.kind == "workbook"
    assert stored.inputs["run_id"] == record.run_id
    assert stored.content_digest == written.content_digest

    # Writing it again is the same artifact, not a second one.
    again = write_workbook(tmp_path / "b.xlsx", result, record,
                           assumptions=ASSUMPTIONS)
    record_workbook(again, registry)
    assert len(registry) == 1

    # The same inputs with different content is a derivation that is not
    # reproducible, and the registry is where that gets noticed.
    drifted = dataclasses.replace(again, content_digest="f" * 32)
    with pytest.raises(ArtifactConflictError):
        record_workbook(drifted, registry)


def test_the_content_digest_is_over_the_cells_not_the_file(tmp_path, run):
    """So it survives an openpyxl upgrade that reorders an XML attribute —
    and so that two workbooks differing only in what was asked for differ."""
    result, record = run
    plain = write_workbook(tmp_path / "a.xlsx", result, record)
    with_detail = write_workbook(tmp_path / "b.xlsx", result, record,
                                 detail=["claims"])
    assert plain.content_digest != with_detail.content_digest
    assert plain.sheets == (SUMMARY_SHEET, AGGREGATE_SHEET, ASSUMPTION_SHEET)
    assert "Detail claims" in with_detail.sheets


def test_the_stochastic_result_type_is_handled_by_the_same_path(stochastic):
    """Guards the assumption the scenario refusals are written against."""
    result, _ = stochastic
    assert isinstance(result, StochasticRunResult)
    assert result.array("fund_eoy").ndim == 3
